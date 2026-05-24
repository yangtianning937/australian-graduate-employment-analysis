"""Extract tidy CSV files from the public QILT GOS-L 2025 workbook.

The original workbook is published by QILT as part of the GOS-L National
Report Tables. This script keeps the R analysis simple by converting selected
report sheets into clean CSV files.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data/raw/GOSL_2025_National_Report_Tables/GOSL_2025_National_Tables_Public.xlsx"
OUT_DIR = ROOT / "data/raw/qilt_extracted"


def parse_number(value):
    if value is None:
        return None
    text = str(value).replace(",", "").strip()
    if text.lower() in {"", "n/a", "na"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_ci_value(value):
    if value is None:
        return (None, None, None)
    text = str(value).strip()
    match = re.match(r"^([0-9,.]+)\s*\(([0-9,.]+),\s*([0-9,.]+)\)$", text)
    if match:
        return tuple(float(item.replace(",", "")) for item in match.groups())
    return (parse_number(value), None, None)


def write_csv(path, fieldnames, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def read_area_summary(wb):
    ws = wb["STMT_UG_ALL_1Y_AREA"]
    rows = []
    for r in range(5, 27):
        field = ws.cell(r, 2).value
        if not field or field in {"Total", "Standard deviation"}:
            continue
        rows.append(
            {
                "source": "QILT GOS-L 2025 National Report Tables",
                "survey_year": 2025,
                "cohort": "Domestic undergraduate",
                "field_of_education": field,
                "short_term_full_time_employment_rate": parse_number(ws.cell(r, 3).value),
                "medium_term_full_time_employment_rate": parse_number(ws.cell(r, 4).value),
                "short_term_overall_employment_rate": parse_number(ws.cell(r, 5).value),
                "medium_term_overall_employment_rate": parse_number(ws.cell(r, 6).value),
                "short_term_labour_force_participation_rate": parse_number(ws.cell(r, 7).value),
                "medium_term_labour_force_participation_rate": parse_number(ws.cell(r, 8).value),
                "short_term_median_salary_aud": parse_number(ws.cell(r, 9).value),
                "medium_term_median_salary_aud": parse_number(ws.cell(r, 10).value),
            }
        )
    return rows


def read_gender_summary(wb):
    ws = wb["STMT_UG_ALL_1Y_ARSX"]
    rows = []
    genders = {
        "Male": {
            "fte_short": 3,
            "fte_medium": 4,
            "oe_short": 7,
            "oe_medium": 8,
            "salary_short": 15,
            "salary_medium": 16,
        },
        "Female": {
            "fte_short": 5,
            "fte_medium": 6,
            "oe_short": 9,
            "oe_medium": 10,
            "salary_short": 17,
            "salary_medium": 18,
        },
    }
    for r in range(5, 27):
        field = ws.cell(r, 2).value
        if not field or field in {"Total", "Standard deviation"}:
            continue
        for gender, cols in genders.items():
            rows.append(
                {
                    "source": "QILT GOS-L 2025 National Report Tables",
                    "survey_year": 2025,
                    "cohort": "Domestic undergraduate",
                    "field_of_education": field,
                    "gender": gender,
                    "short_term_full_time_employment_rate": parse_number(ws.cell(r, cols["fte_short"]).value),
                    "medium_term_full_time_employment_rate": parse_number(ws.cell(r, cols["fte_medium"]).value),
                    "short_term_overall_employment_rate": parse_number(ws.cell(r, cols["oe_short"]).value),
                    "medium_term_overall_employment_rate": parse_number(ws.cell(r, cols["oe_medium"]).value),
                    "short_term_median_salary_aud": parse_number(ws.cell(r, cols["salary_short"]).value),
                    "medium_term_median_salary_aud": parse_number(ws.cell(r, cols["salary_medium"]).value),
                }
            )
    return rows


def read_institution_sheet(wb, sheet_name, value_name):
    ws = wb[sheet_name]
    rows = []
    for r in range(5, 80):
        institution = ws.cell(r, 2).value
        raw_value = ws.cell(r, 3).value
        if not institution or not raw_value:
            continue
        value, ci_low, ci_high = parse_ci_value(raw_value)
        rows.append(
            {
                "source": "QILT GOS-L 2025 National Report Tables",
                "survey_year": 2025,
                "cohort": "Domestic undergraduate universities",
                "institution": institution,
                value_name: value,
                f"{value_name}_ci_low": ci_low,
                f"{value_name}_ci_high": ci_high,
            }
        )
    return rows


def main():
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)

    area_rows = read_area_summary(wb)
    gender_rows = read_gender_summary(wb)
    institution_fte_rows = read_institution_sheet(
        wb, "FTE_UG_UNI_1Y_INST_FIG", "medium_term_full_time_employment_rate"
    )
    institution_salary_rows = read_institution_sheet(
        wb, "SAL_UG_UNI_1Y_INST_FIG", "medium_term_median_salary_aud"
    )

    write_csv(
        OUT_DIR / "qilt_gosl_2025_area_summary.csv",
        list(area_rows[0].keys()),
        area_rows,
    )
    write_csv(
        OUT_DIR / "qilt_gosl_2025_gender_summary.csv",
        list(gender_rows[0].keys()),
        gender_rows,
    )
    write_csv(
        OUT_DIR / "qilt_gosl_2025_institution_fte.csv",
        list(institution_fte_rows[0].keys()),
        institution_fte_rows,
    )
    write_csv(
        OUT_DIR / "qilt_gosl_2025_institution_salary.csv",
        list(institution_salary_rows[0].keys()),
        institution_salary_rows,
    )

    print(f"Extracted {len(area_rows)} area rows")
    print(f"Extracted {len(gender_rows)} gender rows")
    print(f"Extracted {len(institution_fte_rows)} institution employment rows")
    print(f"Extracted {len(institution_salary_rows)} institution salary rows")


if __name__ == "__main__":
    main()
